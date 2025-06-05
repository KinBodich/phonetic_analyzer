import os
from openpyxl import Workbook, load_workbook

class OutputWriter:
    def __init__(self, output_folder):
        self.output_folder = output_folder
        self.vowels = {
            "AA", "AE", "AH", "AO", "AW", "AY", "EH", "ER",
            "EY", "IH", "IY", "OW", "OY", "UH", "UW"
        }

    def write_transcription(self, name, data):
        path = os.path.join(self.output_folder, "Transcribed", f"{name}.txt")
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            f.write(" ".join(" ".join(phonemes) for _, phonemes in data))

    def write_syllables(self, name, data):
        path = os.path.join(self.output_folder, "Syllables", f"{name}.txt")
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            result = []
            for _, syllables in data:
                syll_str = "/".join("-".join(p for p in s) for s in syllables)
                result.append(syll_str)
            f.write(" ".join(result))

    def write_syllables_cvv(self, name, data):
        path = os.path.join(self.output_folder, "SyllablesCVV", f"{name}.txt")
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            output_line = " ".join(
                "/".join("".join("V" if p.rstrip("012") in self.vowels else "C" for p in syll)
                         for syll in syllables)
                for _, syllables in data
            )
            f.write(output_line)

    def write_first_syllables(self, name, data):
        path = os.path.join(self.output_folder, "FirstSyllables", f"{name}.txt")
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            output_line = " ".join(
                "-".join(syllables[0]) if syllables else ""
                for _, syllables in data
            )
            f.write(output_line)

    def write_first_syllables_cvv(self, name, data):
        path = os.path.join(self.output_folder, "FirstSyllablesCVV", f"{name}.txt")
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            output_line = " ".join(
                "".join("V" if p.rstrip("012") in self.vowels else "C" for p in syllables[0])
                if syllables else ""
                for _, syllables in data
            )
            f.write(output_line)

    def write_statistics(self, name, data):
        path = os.path.join(self.output_folder, "Statistics.xlsx")
        if not data:
            return

        # Ключі, які фіксуємо спочатку
        fixed_keys = ["Text", "Length", "PhonemeCount", "SyllablesCount", "AverageSyllables"]

        # Беремо перший рядок, щоб зберегти порядок додаткових ключів
        first_row = data[0]
        additional_keys = [k for k in first_row.keys() if k not in set(fixed_keys)]
        fieldnames = fixed_keys + additional_keys

        os.makedirs(os.path.dirname(path), exist_ok=True)

        # Якщо файл існує, очищуємо його перед перезаписом
        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb.active
            # Видаляємо всі рядки, крім заголовка
            ws.delete_rows(2, ws.max_row)
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(fieldnames)

        # Додаємо всі рядки даних
        for row in data:
            ws.append([row.get(fn, "") for fn in fieldnames])

        # Обчислюємо підсумки
        sum_length = sum(row.get("Length", 0) for row in data)
        sum_phoneme = sum(row.get("PhonemeCount", 0) for row in data)
        sum_syllables = sum(row.get("SyllablesCount", 0) for row in data)
        sum_c = sum(row.get("Total C", 0) for row in data)
        sum_v = sum(row.get("Total V", 0) for row in data)

        # Знаходимо індекси відповідних колонок у fieldnames
        idx_text = fieldnames.index("Text")
        idx_length = fieldnames.index("Length")
        idx_phoneme = fieldnames.index("PhonemeCount")
        idx_syllables = fieldnames.index("SyllablesCount")
        idx_c = fieldnames.index("Total C")
        idx_v = fieldnames.index("Total V")

        # Додаємо порожній рядок перед підсумками
        ws.append([""] * len(fieldnames))

        # Рядок: загальна довжина файлів
        row_length = [""] * len(fieldnames)
        row_length[idx_text] = "Total Length"
        row_length[idx_length] = sum_length
        ws.append(row_length)

        # Рядок: загальна кількість фонем
        row_phoneme = [""] * len(fieldnames)
        row_phoneme[idx_text] = "Total Phonemes"
        row_phoneme[idx_phoneme] = sum_phoneme
        ws.append(row_phoneme)

        # Рядок: загальна кількість складів
        row_syllables = [""] * len(fieldnames)
        row_syllables[idx_text] = "Total Syllables"
        row_syllables[idx_syllables] = sum_syllables
        ws.append(row_syllables)

        # Рядок: сумарна кількість приголосних (C)
        row_c = [""] * len(fieldnames)
        row_c[idx_text] = "Total C"
        row_c[idx_c] = sum_c
        ws.append(row_c)

        # Рядок: сумарна кількість голосних (V)
        row_v = [""] * len(fieldnames)
        row_v[idx_text] = "Total V"
        row_v[idx_v] = sum_v
        ws.append(row_v)

        wb.save(path)

    def write_unknown(self, name, unknown_list):
        path = os.path.join(self.output_folder, "UnknownWords", f"{name}_unknown.txt")
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            for w in unknown_list:
                f.write(w + "\n")
